from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TemplatePreset


def _try_number(s: str):
    """문자열이 순수 숫자면 int/float로 변환, 아니면 원본 반환"""
    s = s.strip()
    if not s:
        return s
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def _escape_excel_string(s: str) -> str:
    """엑셀 문자열 리터럴용 따옴표 이스케이프"""
    return s.replace('"', '""')


# ── 스타일 상수 ──
_header_font = Font(bold=True, size=11, color="FFFFFF")
_header_fill = PatternFill("solid", fgColor="4472C4")
_group_font = Font(bold=True, size=11, color="1F4E79")
_group_fill = PatternFill("solid", fgColor="D6E4F0")
_count_fill = PatternFill("solid", fgColor="F2F2F2")
_pct_fill = PatternFill("solid", fgColor="E2EFDA")
_non_resp_fill = PatternFill("solid", fgColor="FCE4D6")
_red_font = Font(bold=True, size=11, color="FFFFFF")
_red_fill = PatternFill("solid", fgColor="C00000")
_thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_center_align = Alignment(horizontal="center", vertical="center")

RESULT_SHEET = "'결과'"
OVERALL_SHEET = "'전체 통계'"


def _write_stats_formulas(
    ws,
    config: TemplatePreset,
    field_col_map: dict[str, str],
    last_data_row: int,
    file_filter: str | None = None,
) -> int:
    """통계 행을 수식으로 작성 (결과 시트 참조).

    Args:
        ws: 대상 워크시트
        config: 템플릿 설정
        field_col_map: 필드명 → 결과 시트 열 문자
        last_data_row: 결과 시트 마지막 데이터 행 번호 (1-based)
        file_filter: None=전체 통계, 문자열=파일명 필터
    """
    file_col_range = f"{RESULT_SHEET}!$A$2:$A${last_data_row}"

    # B1 = 응답 수
    if file_filter:
        escaped = _escape_excel_string(file_filter)
        total_formula = f'=COUNTIF({file_col_range},"{escaped}")'
    else:
        total_formula = f"=COUNTA({file_col_range})"

    ws["A1"] = "전체 응답 수"
    ws["A1"].font = Font(bold=True, size=11)
    ws["B1"] = total_formula
    ws["B1"].font = Font(bold=True, size=11, color="1F4E79")

    total_ref = "$B$1"
    row = 3  # 1행=제목, 2행=빈 줄

    for field in config.fields:
        if field.is_comment:
            continue

        col_letter = field_col_map.get(field.name)
        if not col_letter:
            continue

        n_opts = len(field.boxes)
        option_labels = []
        for i in range(n_opts):
            if i < len(field.value_map) and field.value_map[i].strip():
                option_labels.append(_try_number(field.value_map[i].strip()))
            else:
                option_labels.append(i + 1)

        if config.reverse_numbering:
            option_labels = list(reversed(option_labels))

        last_col = n_opts + 2  # A:그룹명, B~:옵션, 마지막:미응답
        result_col_range = (
            f"{RESULT_SHEET}!${col_letter}$2:${col_letter}${last_data_row}"
        )

        # ────────── 행 1: 그룹명 + 옵션 레이블 + 미응답 ──────────
        if file_filter:
            # 파일별 시트: 전체 통계에서 참조
            group_formula = f"={OVERALL_SHEET}!$A${row}"
            ws.cell(row=row, column=1, value=group_formula)
            for ci in range(n_opts):
                c = ci + 2
                label_formula = f"={OVERALL_SHEET}!${get_column_letter(c)}${row}"
                ws.cell(row=row, column=c, value=label_formula)
            nr_label_formula = f"={OVERALL_SHEET}!${get_column_letter(last_col)}${row}"
            ws.cell(row=row, column=last_col, value=nr_label_formula)
        else:
            # 전체 통계: 결과 시트 헤더 + 하드코딩 레이블
            group_formula = f"={RESULT_SHEET}!${col_letter}$1"
            ws.cell(row=row, column=1, value=group_formula)
            for ci, label in enumerate(option_labels):
                c = ci + 2
                ws.cell(row=row, column=c, value=label)
            ws.cell(row=row, column=last_col, value="미응답")

        # 스타일
        cell = ws.cell(row=row, column=1)
        cell.font = _group_font
        cell.fill = _group_fill
        cell.border = _thin_border
        cell.alignment = _center_align
        for ci in range(n_opts):
            c = ci + 2
            cell = ws.cell(row=row, column=c)
            cell.font = _header_font
            cell.fill = _header_fill
            cell.border = _thin_border
            cell.alignment = _center_align
        cell = ws.cell(row=row, column=last_col)
        cell.font = _red_font
        cell.fill = _red_fill
        cell.border = _thin_border
        cell.alignment = _center_align

        label_header_row = row

        # ────────── 행 2: 갯수 ──────────
        row += 1
        count_row = row
        ws.cell(row=row, column=1, value="갯수")
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, size=10)
        cell.fill = _count_fill
        cell.border = _thin_border
        cell.alignment = _center_align

        for ci in range(n_opts):
            c = ci + 2
            label_ref = f"${get_column_letter(c)}${label_header_row}"

            if field.allow_duplicates:
                inner = (
                    f'ISNUMBER(SEARCH("," & {label_ref} & ",",'
                    f' "," & ""&{result_col_range} & ","))'
                )
                if file_filter:
                    escaped_fn = _escape_excel_string(file_filter)
                    formula = (
                        f'=SUMPRODUCT(({inner}) * ({file_col_range}="{escaped_fn}"))'
                    )
                else:
                    formula = f"=SUMPRODUCT(--({inner}))"
            else:
                if file_filter:
                    escaped_fn = _escape_excel_string(file_filter)
                    formula = f'=COUNTIFS({result_col_range},{label_ref},{file_col_range},"{escaped_fn}")'
                else:
                    formula = f"=COUNTIF({result_col_range},{label_ref})"

            cell = ws.cell(row=row, column=c, value=formula)
            cell.fill = _count_fill
            cell.border = _thin_border
            cell.alignment = _center_align

        # 미응답 수
        if file_filter:
            escaped_fn = _escape_excel_string(file_filter)
            nr_formula = (
                f'=COUNTIFS({result_col_range},"",{file_col_range},"{escaped_fn}")'
            )
        else:
            nr_formula = f'=COUNTIF({result_col_range},"")'

        cell = ws.cell(row=row, column=last_col, value=nr_formula)
        cell.fill = _non_resp_fill
        cell.border = _thin_border
        cell.alignment = _center_align

        # ────────── 행 3: 응답률 ──────────
        row += 1
        pct_row = row
        ws.cell(row=row, column=1, value="응답률")
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True, size=10)
        cell.fill = _pct_fill
        cell.border = _thin_border
        cell.alignment = _center_align

        for ci in range(n_opts):
            c = ci + 2
            count_ref = f"${get_column_letter(c)}${count_row}"
            formula = f"=IF({total_ref}=0,0,{count_ref}/{total_ref})"
            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = "0.0%"
            cell.fill = _pct_fill
            cell.border = _thin_border
            cell.alignment = _center_align

        # 미응답률
        nr_count_ref = f"${get_column_letter(last_col)}${count_row}"
        formula = f"=IF({total_ref}=0,0,{nr_count_ref}/{total_ref})"
        cell = ws.cell(row=row, column=last_col, value=formula)
        cell.number_format = "0.0%"
        cell.fill = _non_resp_fill
        cell.border = _thin_border
        cell.alignment = _center_align

        row += 2  # 빈 구분 행

    # 열 너비
    ws.column_dimensions["A"].width = 18
    for ci in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    return row


def export_to_excel(
    results: list[dict],
    config: Optional[TemplatePreset] = None,
    out_path: str = "설문결과.xlsx",
) -> bool:
    if not results:
        return False

    try:
        wb = openpyxl.Workbook()

        # ── 1. 결과 시트 ──
        ws_result = wb.active
        ws_result.title = "결과"

        headers = list(results[0].keys())
        ws_result.append(headers)

        for item in results:
            row = [_try_number(str(item.get(header, ""))) for header in headers]
            ws_result.append(row)

        last_data_row = len(results) + 1

        # 필드 → 결과 시트 열 매핑
        field_col_map: dict[str, str] = {}
        if config:
            for idx, h in enumerate(headers):
                col_letter = get_column_letter(idx + 1)
                for field in config.fields:
                    if field.name == h:
                        field_col_map[field.name] = col_letter
                        break

        # ── 2. 전체 통계 시트 ──
        if config and config.fields:
            ws_overall = wb.create_sheet("전체 통계")
            _write_stats_formulas(ws_overall, config, field_col_map, last_data_row)

            # ── 3. 파일별 통계 시트 ──
            fname_set: set[str] = set()
            for item in results:
                fn = str(item.get("파일명", "")).strip()
                if not fn:
                    fn = "기타"
                fname_set.add(fn)

            for fname in sorted(fname_set):
                sheet_name = fname[:31]
                ws_f = wb.create_sheet(sheet_name)
                _write_stats_formulas(
                    ws_f,
                    config,
                    field_col_map,
                    last_data_row,
                    file_filter=fname,
                )

        wb.save(out_path)
        return True

    except Exception as e:
        print(f"엑셀 저장 실패: {e}")
        return False
